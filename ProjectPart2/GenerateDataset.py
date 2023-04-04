from openpyxl import Workbook, load_workbook
import random


class Dataset:
    def __init__(self, dataset_path, test_case, request_num_per_test_case):
        self.dataset_path = dataset_path
        self.test_case = test_case
        self.request_num_per_test_case = request_num_per_test_case

    def generate(self):
        dataset = Workbook()

        dataset['Sheet']['A1'] = self.test_case
        dataset['Sheet']['A2'] = self.request_num_per_test_case

        for i in range(self.test_case):
            sheet_name = 'test_case' + str(i + 1)
            sheet = dataset.create_sheet(sheet_name)

            for j in range(self.request_num_per_test_case):
                disk_position = random.randrange(1, 655356)
                time = random.randrange(1, 100)

                disk_index = 'A' + str(j + 1)
                time_index = 'B' + str(j + 1)

                sheet[disk_index] = disk_position
                sheet[time_index] = time

        dataset.save(self.dataset_path)


def read_dataset(dataset_path):
    all_requests = list()
    dataset = load_workbook(dataset_path)

    test_case = int(dataset['Sheet']['A1'].value)
    request_num_per_test_case = int(dataset['Sheet']['A2'].value)

    for i in range(test_case):
        sheet_name = 'test_case' + str(i + 1)
        sheet = dataset[sheet_name]
        requests = list()

        for j in range(request_num_per_test_case):
            request = list()

            disk_index = 'A' + str(j + 1)
            time_index = 'B' + str(j + 1)

            disk_position = sheet[disk_index].value
            time = sheet[time_index].value

            request.append(float(disk_position))
            request.append(float(time))
            requests.append(request)
        all_requests.append(requests)

    return all_requests
